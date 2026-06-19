"""
Saves leads and reviews to:
  - leads_no_website.xlsx  (two sheets: Leads + Reviews)
  - leads_no_website.csv
  - reviews.csv
"""
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Lead, Review

# ── Column definitions ────────────────────────────────────────────────────────
LEAD_COLS = [
    ("Business ID",       14, "business_id"),
    ("Business Name",     28, "business_name"),
    ("Business Type",     20, "business_type"),
    ("Area",              18, "area"),
    ("Address",           34, "address"),
    ("Phone",             16, "phone"),
    ("Email",             30, "email"),
    ("Rating",            10, "rating"),
    ("Review Count",      14, "review_count"),
    ("Price Range",       12, "price_range"),
    ("Hours",             44, "hours"),
    ("Services",          40, "services"),
    ("About",             55, "about"),
    ("Year Established",  16, "year_established"),
    ("Facebook URL",      38, "facebook_url"),
    ("Instagram URL",     38, "instagram_url"),
    ("Logo URL",          38, "logo_url"),
    ("Maps URL",          38, "maps_url"),
    ("Lead Reason",       28, "lead_reason"),
    ("Score",             10, "score"),
    ("Photos",            34, "photo_dir"),
]

REVIEW_COLS = [
    ("Business ID",    14, "business_id"),
    ("Business Name",  28, "business_name"),
    ("Reviewer",       22, "reviewer"),
    ("Stars",           8, "stars"),
    ("Date",           18, "date"),
    ("Review Text",    72, "text"),
]

# col ranges are 1-indexed, must match LEAD_COLS order above
LEAD_SECTIONS = [
    ("Identity",         "2E75B6",  1,  4),
    ("Contact",          "1F6B35",  5,  7),
    ("Performance",      "6B3A1F",  8, 10),
    ("Business Details", "4B1F6B", 11, 14),
    ("Social Media",     "1F5C6B", 15, 17),
    ("Lead Info",        "6B1F3A", 18, 19),
    ("Ranking",          "3A1F6B", 20, 21),
]


# ── Styling helpers ───────────────────────────────────────────────────────────
def _style_header(cell, bg: str, size: int = 10):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


WRAP_FIELDS  = {"hours", "about", "services", "text"}
LINK_FIELDS  = {"facebook_url", "instagram_url", "logo_url", "maps_url"}
EMAIL_FIELDS = {"email"}

_DATA_FONT  = Font(name="Arial", size=10)
_LINK_FONT  = Font(name="Arial", size=10, color="0563C1", underline="single")
_EMAIL_FONT = Font(name="Arial", size=10, color="0563C1")
_ALT_FILL   = PatternFill("solid", start_color="EEF4FB")
_YELLOW     = PatternFill("solid", start_color="FFF2CC")
_PINK       = PatternFill("solid", start_color="FCE4EC")


def _build_sheet(ws, sections, col_defs, rows):
    # Row 1: section headers
    for label, bg, c_start, c_end in sections:
        ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
        _style_header(ws.cell(row=1, column=c_start, value=label), bg, size=11)
    ws.row_dimensions[1].height = 22

    # Row 2: column headers
    for c, (label, width, _) in enumerate(col_defs, 1):
        _style_header(ws.cell(row=2, column=c, value=label), "1F4E79")
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[2].height = 28

    # Data rows
    for r, obj in enumerate(rows, 3):
        reason   = getattr(obj, "lead_reason", "")
        row_fill = (
            _YELLOW   if "No website" in reason else
            _PINK     if "Outdated"   in reason else
            (_ALT_FILL if r % 2 == 0 else None)
        )
        for c, (_, _, fname) in enumerate(col_defs, 1):
            val  = getattr(obj, fname, "") or ""
            cell = ws.cell(row=r, column=c, value=val)
            wrap = fname in WRAP_FIELDS
            cell.alignment = Alignment(vertical="center", wrap_text=wrap)
            if fname in LINK_FIELDS and val:
                cell.font = _LINK_FONT
            elif fname in EMAIL_FIELDS and val:
                cell.font = _EMAIL_FONT
            else:
                cell.font = _DATA_FONT
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[r].height = 40 if any(
            getattr(obj, f, "") for f in WRAP_FIELDS
        ) else 18

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(col_defs))}2"


# ── Public functions ──────────────────────────────────────────────────────────
def save_excel(leads: list, reviews: list, path: str):
    wb = openpyxl.Workbook()

    ws_leads = wb.active
    ws_leads.title = "Leads"
    _build_sheet(ws_leads, LEAD_SECTIONS, LEAD_COLS, leads)

    ws_rev = wb.create_sheet("Reviews")
    review_sections = [
        ("Reviews — linked to Leads sheet by Business ID", "4B1F6B", 1, len(REVIEW_COLS))
    ]
    _build_sheet(ws_rev, review_sections, REVIEW_COLS, reviews)

    wb.save(path)
    print(f"  Excel -> {path}  ({len(leads)} leads, {len(reviews)} reviews)")


def save_csv(leads: list, reviews: list, leads_path: str, reviews_path: str):
    with open(leads_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([label for label, _, _ in LEAD_COLS])
        for lead in leads:
            w.writerow([getattr(lead, fname, "") or "" for _, _, fname in LEAD_COLS])

    with open(reviews_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([label for label, _, _ in REVIEW_COLS])
        for rev in reviews:
            w.writerow([getattr(rev, fname, "") or "" for _, _, fname in REVIEW_COLS])

    print(f"  CSV   -> {leads_path}")
    print(f"  CSV   -> {reviews_path}")
